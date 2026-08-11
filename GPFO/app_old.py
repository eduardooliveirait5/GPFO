"""
Gestor de Presenças — Formações Online
=======================================
Aplicação com interface gráfica para importar relatórios do Microsoft Teams,
gerir formações e exportar relatórios de aproveitamento em Excel.

Requisitos:
    pip install openpyxl

Uso:
    python app.py

Rev cliente:
    05_2026_01
"""

import os
import re
import sys
import sqlite3
import threading
from datetime import datetime, date
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import warnings
warnings.filterwarnings("ignore")


# ═══════════════════════════════════════════════════════════
#  WIDGETS REUTILIZÁVEIS
# ═══════════════════════════════════════════════════════════

class Label(tk.Label):
    def __init__(self, parent, text="", size=11, bold=False,
                 color=None, **kw):
        color = color or CORES["text"]
        font = ("Segoe UI", size, "bold" if bold else "normal")
        super().__init__(parent, text=text, font=font, fg=color,
                         bg=kw.pop("bg", CORES["bg"]), **kw)


class Entry(tk.Entry):
    def __init__(self, parent, **kw):
        super().__init__(parent,
            font=("Segoe UI", 11),
            bg=CORES["surface2"], fg=CORES["text_bright"],
            insertbackground=CORES["accent"],
            relief="flat", bd=0,
            highlightthickness=1,
            highlightcolor=CORES["accent"],
            highlightbackground=CORES["border"],
            **kw)


class Button(tk.Button):
    def __init__(self, parent, text="", command=None,
                 style="primary", **kw):
        colors = {
            "primary": (CORES["accent"],   CORES["text_bright"]),
            "success": (CORES["success"],  CORES["text_bright"]),
            "danger":  (CORES["danger"],   CORES["text_bright"]),
            "ghost":   (CORES["surface2"], CORES["text"]),
        }
        bg, fg = colors.get(style, colors["primary"])
        super().__init__(parent, text=text, command=command,
            font=("Segoe UI", 10, "bold"),
            bg=bg, fg=fg, activebackground=fg, activeforeground=bg,
            relief="flat", bd=0, cursor="hand2",
            padx=18, pady=8, **kw)


class Card(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent,
            bg=CORES["card"],
            highlightthickness=1,
            highlightbackground=CORES["border"],
            **kw)



class Separator(tk.Frame):
    def __init__(self, parent, **kw):
        kw.setdefault("bg", CORES["border"])
        super().__init__(parent, height=1, **kw)


# Treeview estilizado
def make_tree(parent, cols, heights=None, show="headings"):
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Dark.Treeview",
        background=CORES["surface"],
        foreground=CORES["text"],
        fieldbackground=CORES["surface"],
        rowheight=32,
        font=("Segoe UI", 10),
        borderwidth=0,
        relief="flat")
    style.configure("Dark.Treeview.Heading",
        background=CORES["surface2"],
        foreground=CORES["text_bright"],
        font=("Segoe UI", 10, "bold"),
        relief="flat", borderwidth=0)
    style.map("Dark.Treeview",
        background=[("selected", CORES["accent2"])],
        foreground=[("selected", CORES["text_bright"])])
    style.map("Dark.Treeview.Heading",
        background=[("active", CORES["border"])])

    tree = ttk.Treeview(parent, columns=cols, show=show,
                        style="Dark.Treeview")
    vsb = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)

    # tags de alternância
    tree.tag_configure("odd",  background=CORES["surface"])
    tree.tag_configure("even", background=CORES["surface2"])

    return tree, vsb


# ═══════════════════════════════════════════════════════════
#  JANELA PRINCIPAL
# ═══════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gestor de Presenças — Formações Online do GRUPO XXX")
        self.geometry("1180x720")
        self.minsize(900, 600)
        self.configure(bg=CORES["bg"])

        # Ícone (ignora se não existir)
        try:
            self.iconbitmap("icon.ico")
        except Exception:
            pass

        init_db()
        self._build_ui()
        self.show_page("importar")

    # ── Layout base ────────────────────────────────────────────────

    def _build_ui(self):
        # Sidebar
        self.sidebar = tk.Frame(self, bg=CORES["surface"],
                                width=220, relief="flat")
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        # Logo/título
        tk.Frame(self.sidebar, bg=CORES["accent"], height=4).pack(fill="x")
        logo_frame = tk.Frame(self.sidebar, bg=CORES["surface"], pady=20)
        logo_frame.pack(fill="x")
        Label(logo_frame, "⬡ PRESENÇAS", size=14, bold=True,
              color=CORES["text_bright"], bg=CORES["surface"]).pack()
        Label(logo_frame, "Gestão de Formações", size=9,
              color=CORES["text_dim"], bg=CORES["surface"]).pack()

        Separator(self.sidebar, bg=CORES["border"]).pack(fill="x", padx=16)

        # Botões de navegação
        nav_items = [
            ("importar",  "  ＋  Importar Sessão",     "primary"),
            ("formacoes", "  ◈  Formações",           "ghost"),
            ("relatorio", "  ↓  Exportar Relatório",  "ghost"),
        ]
        self.nav_buttons = {}
        nav_frame = tk.Frame(self.sidebar, bg=CORES["surface"], pady=12)
        nav_frame.pack(fill="x")
        for key, label, _ in nav_items:
            btn = tk.Button(nav_frame, text=label,
                command=lambda k=key: self.show_page(k),
                font=("Segoe UI", 11), bg=CORES["surface"],
                fg=CORES["text"], activebackground=CORES["surface2"],
                activeforeground=CORES["text_bright"],
                relief="flat", bd=0, cursor="hand2",
                anchor="w", padx=20, pady=10, width=22)
            btn.pack(fill="x")
            self.nav_buttons[key] = btn

        # Rodapé sidebar
        tk.Frame(self.sidebar, bg=CORES["surface"]).pack(fill="both", expand=True)
        Label(self.sidebar, "v2.3  •  SQLite + Excel - Eduardo Oliveira", size=8,
              color=CORES["text_dim"], bg=CORES["surface"]).pack(pady=10)

        # Área de conteúdo
        self.content = tk.Frame(self, bg=CORES["bg"])
        self.content.pack(side="left", fill="both", expand=True)

        # Páginas
        self.pages: dict[str, tk.Frame] = {}
        for PageClass, key in [
            (PageImportar,  "importar"),
            (PageFormacoes, "formacoes"),
            (PageRelatorio, "relatorio"),
        ]:
            page = PageClass(self.content, self)
            self.pages[key] = page

    def show_page(self, key: str):
        for k, btn in self.nav_buttons.items():
            if k == key:
                btn.configure(bg=CORES["surface2"],
                              fg=CORES["text_bright"])
            else:
                btn.configure(bg=CORES["surface"],
                              fg=CORES["text"])

        for k, page in self.pages.items():
            if k == key:
                page.place(relx=0, rely=0, relwidth=1, relheight=1)
                if hasattr(page, "on_show"):
                    page.on_show()
            else:
                page.place_forget()

    def notify(self, msg: str, kind="info"):
        """Toast de notificação no topo."""
        cores_kind = {"info": CORES["accent"], "ok": CORES["success"],
                      "err": CORES["danger"]}
        cor = cores_kind.get(kind, CORES["accent"])
        toast = tk.Frame(self.content, bg=cor, pady=8, padx=16)
        toast.place(relx=0.5, y=16, anchor="n")
        tk.Label(toast, text=msg, font=("Segoe UI", 10, "bold"),
                 bg=cor, fg="white").pack()
        self.after(3200, toast.destroy)


# ═══════════════════════════════════════════════════════════
#  PÁGINA: IMPORTAR
# ═══════════════════════════════════════════════════════════

class PageImportar(tk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent, bg=CORES["bg"])
        self.app = app
        self._ficheiro = tk.StringVar()
        self._codigo   = tk.StringVar()
        self._nome_form = tk.StringVar()
        self._h_ini    = tk.StringVar(value="")
        self._h_fim    = tk.StringVar(value="")
        self._usar_janela = tk.BooleanVar(value=False)
        self._build()

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=CORES["bg"], pady=28, padx=36)
        hdr.pack(fill="x")
        Label(hdr, "Importar Sessão", size=20, bold=True,
              color=CORES["text_bright"], bg=CORES["bg"]).pack(anchor="w")
        Label(hdr, "Importa o relatório de presença exportado pelo Microsoft Teams",
              size=11, color=CORES["text_dim"], bg=CORES["bg"]).pack(anchor="w")

        # Conteúdo
        body = tk.Frame(self, bg=CORES["bg"], padx=36)
        body.pack(fill="both", expand=True)

        # ── Card: Ficheiro ───────────────────────────────────────────
        card1 = Card(body)
        card1.pack(fill="x", pady=(0, 14))
        inner1 = tk.Frame(card1, bg=CORES["card"], padx=24, pady=20)
        inner1.pack(fill="x")

        Label(inner1, "1.  Ficheiro Excel do Teams", size=12, bold=True,
              color=CORES["text_bright"], bg=CORES["card"]).pack(anchor="w")
        Label(inner1, "Selecione o ficheiro .xlsx exportado após reunião do Teams",
              size=10, color=CORES["text_dim"], bg=CORES["card"]).pack(anchor="w", pady=(2, 12))

        row_file = tk.Frame(inner1, bg=CORES["card"])
        row_file.pack(fill="x")
        ent = Entry(row_file, textvariable=self._ficheiro, width=52)
        ent.pack(side="left", ipady=6, padx=(0, 10))
        Button(row_file, "📂  Procurar…",
               command=self._browse, style="ghost").pack(side="left")

        # ── Card: Formação ───────────────────────────────────────────
        card2 = Card(body)
        card2.pack(fill="x", pady=(0, 14))
        inner2 = tk.Frame(card2, bg=CORES["card"], padx=24, pady=20)
        inner2.pack(fill="x")

        Label(inner2, "2.  Formação", size=12, bold=True,
              color=CORES["text_bright"], bg=CORES["card"]).pack(anchor="w")
        Label(inner2, "Código único para agrupar várias aulas da mesma formação",
              size=10, color=CORES["text_dim"], bg=CORES["card"]).pack(anchor="w", pady=(2, 12))

        row2 = tk.Frame(inner2, bg=CORES["card"])
        row2.pack(fill="x")

        tk.Frame(row2, bg=CORES["card"]).pack(side="left")
        col_cod = tk.Frame(row2, bg=CORES["card"])
        col_cod.pack(side="left", padx=(0, 24))
        Label(col_cod, "Inserir Código", size=9, color=CORES["text_dim"],
              bg=CORES["card"]).pack(anchor="w")
        Entry(col_cod, textvariable=self._codigo, width=18).pack(ipady=6)

        col_nom = tk.Frame(row2, bg=CORES["card"])
        col_nom.pack(side="left")
        Label(col_nom, "Nome da Formação (opcional)", size=9,
              color=CORES["text_dim"], bg=CORES["card"]).pack(anchor="w")
        Entry(col_nom, textvariable=self._nome_form, width=36).pack(ipady=6)

        # ── Card: Janela Horária ─────────────────────────────────────
        card3 = Card(body)
        card3.pack(fill="x", pady=(0, 14))
        inner3 = tk.Frame(card3, bg=CORES["card"], padx=24, pady=20)
        inner3.pack(fill="x")

        Label(inner3, "3.  Janela Horária (opcional)", size=12, bold=True,
              color=CORES["text_bright"], bg=CORES["card"]).pack(anchor="w")
        Label(inner3,
              "Define o período da formação — só é contabilizado o tempo dentro deste intervalo",
              size=10, color=CORES["text_dim"], bg=CORES["card"]).pack(anchor="w", pady=(2, 12))

        chk_frame = tk.Frame(inner3, bg=CORES["card"])
        chk_frame.pack(anchor="w", pady=(0, 10))
        tk.Checkbutton(chk_frame, text="Definir horário da formação",
            variable=self._usar_janela, command=self._toggle_janela,
            font=("Segoe UI", 10), bg=CORES["card"],
            fg=CORES["text"], selectcolor=CORES["surface2"],
            activebackground=CORES["card"], activeforeground=CORES["text"],
            cursor="hand2").pack(side="left")

        self._row_janela = tk.Frame(inner3, bg=CORES["card"])
        self._row_janela.pack(anchor="w")

        col_ini = tk.Frame(self._row_janela, bg=CORES["card"])
        col_ini.pack(side="left", padx=(0, 20))
        Label(col_ini, "Hora de início  (HH:MM)", size=9,
              color=CORES["text_dim"], bg=CORES["card"]).pack(anchor="w")
        self._ent_ini = Entry(col_ini, textvariable=self._h_ini, width=10)
        self._ent_ini.pack(ipady=6)

        col_fim = tk.Frame(self._row_janela, bg=CORES["card"])
        col_fim.pack(side="left")
        Label(col_fim, "Hora de fim  (HH:MM)", size=9,
              color=CORES["text_dim"], bg=CORES["card"]).pack(anchor="w")
        self._ent_fim = Entry(col_fim, textvariable=self._h_fim, width=10)
        self._ent_fim.pack(ipady=6)

        self._toggle_janela()

        # ── Botão importar ───────────────────────────────────────────
        row_btn = tk.Frame(body, bg=CORES["bg"], pady=10)
        row_btn.pack(fill="x")
        Button(row_btn, "  ↑  Importar Sessão ",
               command=self._importar, style="success").pack(side="left")

        # Status
        self._status = Label(row_btn, "", size=14,
                             color=CORES["text_dim"], bg=CORES["bg"])
        self._status.pack(side="left", padx=16)

    def _toggle_janela(self):
        state = "normal" if self._usar_janela.get() else "disabled"
        fg_d = CORES["text_dim"] if state == "disabled" else CORES["text"]
        self._ent_ini.configure(state=state,
            highlightbackground=CORES["border"] if state == "disabled" else CORES["accent"])
        self._ent_fim.configure(state=state,
            highlightbackground=CORES["border"] if state == "disabled" else CORES["accent"])

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Selecionar relatório Teams",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
        if path:
            self._ficheiro.set(path)

    def _importar(self):
        ficheiro = self._ficheiro.get().strip()
        codigo   = self._codigo.get().strip().upper()
        nome     = self._nome_form.get().strip()

        if not ficheiro:
            self.app.notify("Seleciona um ficheiro Excel.", "err")
            return
        if not codigo:
            self.app.notify("Insere o código da formação.", "err")
            return
        if not os.path.isfile(ficheiro):
            self.app.notify("Ficheiro não encontrado.", "err")
            return

        h_ini = self._h_ini.get().strip() if self._usar_janela.get() else None
        h_fim = self._h_fim.get().strip() if self._usar_janela.get() else None

        if self._usar_janela.get():
            for v in [h_ini, h_fim]:
                if not re.match(r"^\d{2}:\d{2}$", v or ""):
                    self.app.notify("Formato de hora inválido — use HH:MM", "err")
                    return

        self._status.configure(text="A importar…")
        self.update_idletasks()

        def run():
            try:
                sessao_id = importar_sessao(ficheiro, codigo, h_ini, h_fim)
                # Actualiza nome da formação se fornecido
                if nome:
                    with get_conn() as conn:
                        conn.execute(
                            "UPDATE formacoes SET nome=? WHERE codigo=?",
                            (nome, codigo))
                        conn.commit()
                self.after(0, lambda: self._success(sessao_id))
            except Exception as e:
                self.after(0, lambda: self._error(str(e)))

        threading.Thread(target=run, daemon=True).start()

    def _success(self, sessao_id):
        self._status.configure(text=f"✓ Sessão importada (id={sessao_id})")
        self.app.notify("Aula importada com sucesso!", "ok")
        self._ficheiro.set("")

    def _error(self, msg):
        self._status.configure(text="")
        self.app.notify(f"Erro: {msg}", "err")


# ═══════════════════════════════════════════════════════════
#  PÁGINA: FORMAÇÕES
# ═══════════════════════════════════════════════════════════

class PageFormacoes(tk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent, bg=CORES["bg"])
        self.app = app
        self._sel_form = None
        self._build()

    def on_show(self):
        self._load_formacoes()

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=CORES["bg"], pady=28, padx=36)
        hdr.pack(fill="x")
        Label(hdr, "Formações", size=20, bold=True,
              color=CORES["text_bright"], bg=CORES["bg"]).pack(anchor="w")
        Label(hdr, "Consulta as formações e sessões importadas",
              size=11, color=CORES["text_dim"], bg=CORES["bg"]).pack(anchor="w")

        body = tk.Frame(self, bg=CORES["bg"], padx=36)
        body.pack(fill="both", expand=True)

        # Split: lista formações (esq) + detalhe (dir)
        paned = tk.PanedWindow(body, orient="horizontal",
                               bg=CORES["bg"], sashwidth=6,
                               sashrelief="flat", bd=0)
        paned.pack(fill="both", expand=True)

        # ── Painel esquerdo: lista ───────────────────────────────────
        left = tk.Frame(paned, bg=CORES["bg"])
        paned.add(left, minsize=260)

        Label(left, "FORMAÇÕES", size=9, bold=True,
              color=CORES["text_dim"], bg=CORES["bg"]).pack(anchor="w", pady=(0, 6))

        tree_frame = tk.Frame(left, bg=CORES["bg"])
        tree_frame.pack(fill="both", expand=True)

        self.tree_form, vsb = make_tree(
            tree_frame, ["codigo", "nome", "sessoes"], show="headings")
        self.tree_form.heading("codigo",  text="Código")
        self.tree_form.heading("nome",    text="Nome")
        self.tree_form.heading("sessoes", text="Sessões")
        self.tree_form.column("codigo",  width=100, anchor="center")
        self.tree_form.column("nome",    width=140)
        self.tree_form.column("sessoes", width=60, anchor="center")
        self.tree_form.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree_form.bind("<<TreeviewSelect>>", self._on_form_select)

        # Botão eliminar formação
        btn_row = tk.Frame(left, bg=CORES["bg"], pady=8)
        btn_row.pack(fill="x")
        Button(btn_row, "Eliminar Formação",
               command=self._eliminar_formacao, style="danger").pack(side="left")

        # ── Painel direito: detalhe ──────────────────────────────────
        right = tk.Frame(paned, bg=CORES["bg"], padx=16)
        paned.add(right, minsize=360)

        self._lbl_form = Label(right, "Seleciona uma formação",
                               size=13, bold=True,
                               color=CORES["text_bright"], bg=CORES["bg"])
        self._lbl_form.pack(anchor="w", pady=(0, 4))

        Label(right, "SESSÕES IMPORTADAS", size=9, bold=True,
              color=CORES["text_dim"], bg=CORES["bg"]).pack(anchor="w", pady=(12, 6))

        sess_frame = tk.Frame(right, bg=CORES["bg"])
        sess_frame.pack(fill="both", expand=True)

        self.tree_sess, vsb2 = make_tree(
            sess_frame, ["data", "titulo", "janela", "dur", "alunos"], show="headings")
        self.tree_sess.heading("data",   text="Data")
        self.tree_sess.heading("titulo", text="Sessão")
        self.tree_sess.heading("janela", text="Janela")
        self.tree_sess.heading("dur",    text="Dur.(min)")
        self.tree_sess.heading("alunos", text="Alunos")
        self.tree_sess.column("data",   width=90, anchor="center")
        self.tree_sess.column("titulo", width=200)
        self.tree_sess.column("janela", width=110, anchor="center")
        self.tree_sess.column("dur",    width=80, anchor="center")
        self.tree_sess.column("alunos", width=60, anchor="center")
        self.tree_sess.pack(side="left", fill="both", expand=True)
        vsb2.pack(side="right", fill="y")

        btn_row2 = tk.Frame(right, bg=CORES["bg"], pady=8)
        btn_row2.pack(fill="x")
        Button(btn_row2, "Eliminar Sessão Selecionada",
               command=self._eliminar_sessao, style="danger").pack(side="left")

    def _load_formacoes(self):
        for i in self.tree_form.get_children():
            self.tree_form.delete(i)
        with get_conn() as conn:
            rows = conn.execute("""
                SELECT f.codigo, f.nome,
                       COUNT(s.id) AS sessoes
                FROM formacoes f
                LEFT JOIN sessoes s ON f.codigo = s.formacao_codigo
                GROUP BY f.codigo
                ORDER BY f.codigo
            """).fetchall()
        for i, r in enumerate(rows):
            tag = "odd" if i % 2 else "even"
            self.tree_form.insert("", "end",
                values=(r["codigo"], r["nome"], r["sessoes"]),
                iid=r["codigo"], tags=(tag,))

    def _on_form_select(self, _=None):
        sel = self.tree_form.selection()
        if not sel:
            return
        codigo = sel[0]
        self._sel_form = codigo
        self._lbl_form.configure(text=f"Formação:  {codigo}")
        self._load_sessoes(codigo)

    def _load_sessoes(self, codigo: str):
        for i in self.tree_sess.get_children():
            self.tree_sess.delete(i)
        with get_conn() as conn:
            rows = conn.execute("""
                SELECT s.id, s.data, s.titulo,
                       s.hora_inicio_janela, s.hora_fim_janela,
                       s.duracao_sessao_min,
                       COUNT(p.id) AS alunos
                FROM sessoes s
                LEFT JOIN presencas p ON p.sessao_id = s.id
                    AND p.funcao != 'Organizador'
                WHERE s.formacao_codigo = ?
                GROUP BY s.id
                ORDER BY s.data
            """, (codigo,)).fetchall()
        for i, r in enumerate(rows):
            janela = (f"{r['hora_inicio_janela']}–{r['hora_fim_janela']}"
                      if r["hora_inicio_janela"] else "Original")
            tag = "odd" if i % 2 else "even"
            self.tree_sess.insert("", "end", iid=str(r["id"]),
                values=(r["data"], r["titulo"], janela,
                        int(r["duracao_sessao_min"]), r["alunos"]),
                tags=(tag,))

    def _eliminar_formacao(self):
        if not self._sel_form:
            return
        if not messagebox.askyesno(
                "Confirmar",
                f"Eliminar formação '{self._sel_form}' e todas as suas sessões?",
                icon="warning"):
            return
        with get_conn() as conn:
            sess_ids = [r[0] for r in conn.execute(
                "SELECT id FROM sessoes WHERE formacao_codigo=?",
                (self._sel_form,)).fetchall()]
            for sid in sess_ids:
                conn.execute("DELETE FROM presencas WHERE sessao_id=?", (sid,))
            conn.execute("DELETE FROM sessoes WHERE formacao_codigo=?",
                         (self._sel_form,))
            conn.execute("DELETE FROM formacoes WHERE codigo=?",
                         (self._sel_form,))
            conn.commit()
        self._sel_form = None
        self._lbl_form.configure(text="Seleciona uma formação")
        for i in self.tree_sess.get_children():
            self.tree_sess.delete(i)
        self._load_formacoes()
        self.app.notify("Formação eliminada.", "ok")

    def _eliminar_sessao(self):
        sel = self.tree_sess.selection()
        if not sel:
            self.app.notify("Seleciona uma sessão.", "err")
            return
        sid = int(sel[0])
        if not messagebox.askyesno(
                "Confirmar", "Eliminar esta sessão e as suas presenças?",
                icon="warning"):
            return
        with get_conn() as conn:
            conn.execute("DELETE FROM presencas WHERE sessao_id=?", (sid,))
            conn.execute("DELETE FROM sessoes WHERE id=?", (sid,))
            conn.commit()
        self._load_sessoes(self._sel_form)
        self._load_formacoes()
        self.app.notify("Sessão eliminada.", "ok")


# ═══════════════════════════════════════════════════════════
#  PÁGINA: EXPORTAR RELATÓRIO
# ═══════════════════════════════════════════════════════════

class PageRelatorio(tk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent, bg=CORES["bg"])
        self.app = app
        self._sel = tk.StringVar()
        self._build()

    def on_show(self):
        self._refresh()

    def _build(self):
        hdr = tk.Frame(self, bg=CORES["bg"], pady=28, padx=36)
        hdr.pack(fill="x")
        Label(hdr, "Exportar Relatório", size=20, bold=True,
              color=CORES["text_bright"], bg=CORES["bg"]).pack(anchor="w")
        Label(hdr, "Gera o ficheiro Excel com o aproveitamento completo da formação",
              size=11, color=CORES["text_dim"], bg=CORES["bg"]).pack(anchor="w")

        body = tk.Frame(self, bg=CORES["bg"], padx=36)
        body.pack(fill="both", expand=True)

        # Card seleção
        card = Card(body)
        card.pack(fill="x", pady=(0, 20))
        inner = tk.Frame(card, bg=CORES["card"], padx=24, pady=24)
        inner.pack(fill="x")

        Label(inner, "Formação a exportar", size=12, bold=True,
              color=CORES["text_bright"], bg=CORES["card"]).pack(anchor="w")
        Label(inner, "Seleciona a formação para gerar o relatório Excel",
              size=10, color=CORES["text_dim"], bg=CORES["card"]).pack(anchor="w", pady=(2, 14))

        row = tk.Frame(inner, bg=CORES["card"])
        row.pack(fill="x")

        self._combo = ttk.Combobox(row, textvariable=self._sel,
            font=("Segoe UI", 11), state="readonly", width=36)
        style = ttk.Style()
        style.configure("TCombobox",
            fieldbackground=CORES["surface2"],
            background=CORES["surface2"],
            foreground=CORES["text_bright"],
            selectbackground=CORES["accent2"])
        self._combo.pack(side="left", ipady=6, padx=(0, 14))
        Button(row, "↻  Actualizar",
               command=self._refresh, style="ghost").pack(side="left")

        Separator(inner, bg=CORES["border"]).pack(fill="x", pady=20)

        # Preview stats
        self._stats = tk.Frame(inner, bg=CORES["card"])
        self._stats.pack(fill="x", pady=(0, 20))
        self._build_stats_empty()

        # Botão exportar
        Button(inner, "  ↓  Exportar para Excel  ",
               command=self._exportar, style="success").pack(anchor="w")

        self._combo.bind("<<ComboboxSelected>>", lambda _: self._show_stats())

    def _build_stats_empty(self):
        for w in self._stats.winfo_children():
            w.destroy()
        Label(self._stats, "Seleciona uma formação para ver o resumo",
              size=10, color=CORES["text_dim"], bg=CORES["card"]).pack(anchor="w")

    def _refresh(self):
        with get_conn() as conn:
            rows = conn.execute("""
                SELECT f.codigo, f.nome, COUNT(s.id) AS sessoes
                FROM formacoes f
                LEFT JOIN sessoes s ON f.codigo = s.formacao_codigo
                GROUP BY f.codigo ORDER BY f.codigo
            """).fetchall()
        values = [f"{r['codigo']}  —  {r['nome']}  ({r['sessoes']} sessões)"
                  for r in rows]
        self._combo_data = {v: r["codigo"] for v, r in zip(values, rows)}
        self._combo["values"] = values
        if values and not self._sel.get():
            self._combo.current(0)
            self._show_stats()

    def _get_codigo(self) -> str | None:
        sel = self._sel.get()
        return self._combo_data.get(sel)

    def _show_stats(self):
        for w in self._stats.winfo_children():
            w.destroy()
        codigo = self._get_codigo()
        if not codigo:
            return
        with get_conn() as conn:
            sessoes_n = conn.execute(
                "SELECT COUNT(*) FROM sessoes WHERE formacao_codigo=?",
                (codigo,)).fetchone()[0]
            alunos_n = conn.execute("""
                SELECT COUNT(DISTINCT p.email) FROM presencas p
                JOIN sessoes s ON p.sessao_id = s.id
                WHERE s.formacao_codigo=? AND p.funcao != 'Organizador'
            """, (codigo,)).fetchone()[0]
            total_min = conn.execute("""
                SELECT SUM(duracao_sessao_min) FROM sessoes
                WHERE formacao_codigo=?
            """, (codigo,)).fetchone()[0] or 0
            aprov_n = conn.execute("""
                SELECT COUNT(DISTINCT p.email)
                FROM presencas p
                JOIN sessoes s ON p.sessao_id = s.id
                WHERE s.formacao_codigo = ?
                  AND p.funcao != 'Organizador'
                GROUP BY p.email
                HAVING SUM(p.duracao_min) * 100.0 / ? >= 75
            """, (codigo, total_min)).fetchall()

        horas = f"{int(total_min//60)}h {int(total_min%60):02d}min"

        stats = [
            ("Sessões",     str(sessoes_n), CORES["accent"]),
            ("Alunos",      str(alunos_n),  CORES["accent2"]),
            ("Carga Total", horas,           CORES["warning"]),
            ("Aprovados",   str(len(aprov_n)), CORES["success"]),
        ]
        for label, val, cor in stats:
            box = tk.Frame(self._stats, bg=CORES["surface"],
                           highlightthickness=1,
                           highlightbackground=CORES["border"],
                           padx=20, pady=14)
            box.pack(side="left", padx=(0, 14))
            tk.Label(box, text=val, font=("Segoe UI", 22, "bold"),
                     fg=cor, bg=CORES["surface"]).pack()
            tk.Label(box, text=label, font=("Segoe UI", 9),
                     fg=CORES["text_dim"], bg=CORES["surface"]).pack()

    def _exportar(self):
        codigo = self._get_codigo()
        if not codigo:
            self.app.notify("Seleciona uma formação.", "err")
            return
        path = filedialog.asksaveasfilename(
            title="Guardar relatório Excel",
            defaultextension=".xlsx",
            initialfile=f"relatorio_{codigo}.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        def run():
            try:
                exportar_formacao_excel(codigo, path)
                self.after(0, lambda: self.app.notify(
                    f"Relatório exportado com sucesso!", "ok"))
            except Exception as e:
                self.after(0, lambda: self.app.notify(
                    f"Erro ao exportar: {e}", "err"))

        threading.Thread(target=run, daemon=True).start()
        self.app.notify("A gerar relatório…", "info")


# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = App()
    app.mainloop()
