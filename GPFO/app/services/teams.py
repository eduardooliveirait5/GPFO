# ═══════════════════════════════════════════════════════════
#  LÓGICA DO TEAMS
#  Toda a lógica de processamento do Microsoft Teams.
# ═══════════════════════════════════════════════════════════
import os
import re
import shutil

from app.config import UPLOAD_DIR
from datetime import datetime
from openpyxl import load_workbook
from app.database import get_conn
from app.logger import logger
from tempfile import NamedTemporaryFile


def parse_duracao_teams(texto: str) -> float:
    if not texto or str(texto).strip() == "":
        return 0.0
    texto = str(texto).strip()
    h = re.search(r"(\d+)\s*h",   texto)
    m = re.search(r"(\d+)\s*min", texto)
    s = re.search(r"(\d+)\s*s",   texto)
    return round((int(h.group(1)) if h else 0) * 60 +
                 (int(m.group(1)) if m else 0) +
                 (int(s.group(1)) if s else 0) / 60, 2)


def parse_datetime_teams(texto) -> datetime | None:
    if not texto:
        return None
    for fmt in ("%m/%d/%y, %I:%M:%S %p", "%d/%m/%Y, %H:%M:%S",
                "%m/%d/%Y, %I:%M:%S %p"):
        try:
            return datetime.strptime(str(texto).strip(), fmt)
        except ValueError:
            continue
    return None


def limpar_nome(nome: str) -> str:
    match = re.search(r"-\s+([A-Za-zÀ-ÿ].+)$", nome.strip())
    nome_limpo = match.group(1).strip() if match else nome.strip()
    return nome_limpo.title()


def calcular_sobreposicao(entrada, saida, j_ini, j_fim) -> float:
    ini = max(entrada, j_ini)
    fim = min(saida,   j_fim)
    if fim <= ini:
        return 0.0
    return round((fim - ini).total_seconds() / 60, 2)


def ler_teams_excel(caminho: str, hora_inicio: str = None, hora_fim: str = None):
    """
    Lê o Excel do Teams e devolve (meta, participantes).
    meta: dict com titulo, data, duracao_min
    participantes: list of dicts
    """
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))

    # Metadados
    meta = {"titulo": "", "data": "", "hora_inicio": None,
            "hora_fim": None, "duracao_min": 0.0}

    for linha in linhas:
        if not linha[0]:
            continue
        chave = str(linha[0]).strip()
        if chave == "Título da reunião":
            meta["titulo"] = str(linha[1]).strip()
        elif chave == "Hora de início":
            dt = parse_datetime_teams(linha[1])
            if dt:
                meta["hora_inicio"] = dt
                meta["data"] = dt.strftime("%Y-%m-%d")
        elif chave == "Hora de fim":
            meta["hora_fim"] = parse_datetime_teams(linha[1])
        elif chave == "Duração da reunião":
            meta["duracao_min"] = parse_duracao_teams(linha[1])

    if meta["duracao_min"] == 0 and meta["hora_inicio"] and meta["hora_fim"]:
        delta = meta["hora_fim"] - meta["hora_inicio"]
        meta["duracao_min"] = round(delta.total_seconds() / 60, 2)

    # Janela horária manual
    usar_janela = False
    j_ini = j_fim = None
    if hora_inicio and hora_fim and meta["hora_inicio"]:
        try:
            data_sessao = meta["hora_inicio"].date()
            h_ini = datetime.strptime(hora_inicio, "%H:%M").time()
            h_fim = datetime.strptime(hora_fim,   "%H:%M").time()
            j_ini = datetime.combine(data_sessao, h_ini)
            j_fim = datetime.combine(data_sessao, h_fim)
            meta["duracao_min"] = round(
                (j_fim - j_ini).total_seconds() / 60, 2)
            usar_janela = True
        except ValueError:
            pass

    # Localiza tabela
    idx = None
    for i, linha in enumerate(linhas):
        if linha[0] == "Nome" and linha[1] == "Hora de Entrada":
            idx = i
            break
    if idx is None:
        raise ValueError("Tabela de participantes não encontrada no ficheiro.")

    # Lê registros
    raw = []
    for linha in linhas[idx + 1:]:
        if not linha[0]:
            break
        nome_raw    = str(linha[0]).strip()
        entrada_dt  = parse_datetime_teams(linha[1])
        saida_dt    = parse_datetime_teams(linha[2])
        duracao_str = str(linha[3]).strip() if linha[3] else ""
        email       = str(linha[4]).strip().lower() if linha[4] else ""
        funcao      = str(linha[5]).strip() if linha[5] else ""

        if funcao.lower() == "organizador":
            continue

        if usar_janela and entrada_dt and saida_dt:
            dur = calcular_sobreposicao(entrada_dt, saida_dt, j_ini, j_fim)
        else:
            dur = parse_duracao_teams(duracao_str)

        raw.append({
            "nome":        limpar_nome(nome_raw),
            "email":       email,
            "funcao":      funcao,
            "duracao_min": dur,
        })

    # Agrega por email
    agg = {}
    for r in raw:
        key = r["email"]
        if key in agg:
            agg[key]["duracao_min"]    += r["duracao_min"]
            agg[key]["num_reconexoes"] += 1
        else:
            agg[key] = {**r, "num_reconexoes": 1}

    wb.close()

    return meta, list(agg.values())


def importar_sessao(
    caminho: str,
    formacao_codigo: str,
    hora_inicio: str = None,
    hora_fim: str = None
) -> int:
    """
    Importa uma sessão do Teams para a base de dados.

    Devolve o ID da sessão criada.
    """

    meta, participantes = ler_teams_excel(
        caminho,
        hora_inicio,
        hora_fim
    )

    with get_conn() as conn:

        # Garante que a formação existe
        conn.execute(
            """
            INSERT OR IGNORE INTO formacoes
            (
                codigo,
                nome,
                criada_em
            )
            VALUES (?,?,?)
            """,
            (
                formacao_codigo,
                formacao_codigo,
                datetime.now().isoformat()
            )
        )

        # Verifica duplicação
        dup = conn.execute(
            """
            SELECT id
            FROM sessoes
            WHERE formacao_codigo=?
              AND titulo=?
              AND data=?
            """,
            (
                formacao_codigo,
                meta["titulo"],
                meta["data"]
            )
        ).fetchone()

        if dup:
            raise ValueError(
                f"Sessão '{meta['titulo']}' de {meta['data']} já importada nesta formação."
            )

        # Cria a sessão
        cur = conn.execute(
            """
            INSERT INTO sessoes
            (
                formacao_codigo,
                titulo,
                data,
                hora_inicio_janela,
                hora_fim_janela,
                duracao_sessao_min,
                ficheiro_original,
                importada_em
            )
            VALUES (?,?,?,?,?,?,?,?)
            """,
            (
                formacao_codigo,
                meta["titulo"],
                meta["data"],
                hora_inicio,
                hora_fim,
                meta["duracao_min"],
                os.path.basename(caminho),
                datetime.now().isoformat()
            )
        )

        sessao_id = cur.lastrowid

        # Guarda os participantes
        for participante in participantes:

            conn.execute(
                """
                INSERT INTO presencas
                (
                    sessao_id,
                    email,
                    nome,
                    funcao,
                    duracao_min,
                    num_reconexoes
                )
                VALUES (?,?,?,?,?,?)
                """,
                (
                    sessao_id,
                    participante["email"],
                    participante["nome"],
                    participante["funcao"],
                    participante["duracao_min"],
                    participante["num_reconexoes"]
                )
            )

        conn.commit()

    return sessao_id

async def importar_excel(
    upload_file,
    codigo,
    nome,
    hora_inicio,
    hora_fim
):
    """
    Recebe um UploadFile do FastAPI e utiliza a função
    importar_sessao() já existente.
    """
    logger.info(f"Iniciando importação da formação {codigo}")

    temp = NamedTemporaryFile(
        dir=UPLOAD_DIR,
        suffix=".xlsx",
        delete=False
    )

    try:

        with temp as buffer:

            shutil.copyfileobj(
                upload_file.file,
                buffer
            )

        sessao = importar_sessao(
            temp.name,
            codigo,
            hora_inicio,
            hora_fim
        )
        logger.info(f"Sessão importada com sucesso. ID={sessao}")
        
        if nome.strip():

            with get_conn() as conn:

                conn.execute(
                    """
                    UPDATE formacoes
                    SET nome=?
                    WHERE codigo=?
                    """,
                    (
                        nome.strip(),
                        codigo
                    )
                )

                conn.commit()

                logger.info(f"Nome da formação atualizado para '{nome}'")

        return sessao

    except Exception:

        logger.exception("Erro durante a importação")

        raise

    finally:

        upload_file.file.close()

        if os.path.exists(temp.name):

            try:
                os.remove(temp.name)
                logger.info("Arquivo temporário removido")
                
            except PermissionError:
                pass