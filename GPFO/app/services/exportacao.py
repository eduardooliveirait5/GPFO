# ═══════════════════════════════════════════════════════════
#  EXPORTAÇÃO EXCEL - Formatação das planilhas
# ═══════════════════════════════════════════════════════════

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter
from app.config import EXPORT_DIR
from openpyxl.formatting.rule import (
    DataBarRule,
    ColorScaleRule,
)
from app.database import get_conn

COR_H   = "1F3864"
COR_H2  = "2E75B6"
COR_OK  = "C6EFCE"
COR_WRN = "FFEB9C"
COR_ERR = "FFC7CE"
COR_A   = "EEF3FB"
COR_B   = "FFFFFF"

def _bd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _c(ws, r, c, v, bold=False, center=False, fundo=None,
       fonte="1F3864", fmt=None, size=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=bold, color=fonte, size=size)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True)
    if fundo:
        cell.fill = PatternFill("solid", start_color=fundo)
    cell.border = _bd()
    if fmt:
        cell.number_format = fmt
    return cell


def _hdr(ws, ln, cols):
    for i, t in enumerate(cols, 1):
        _c(ws, ln, i, t, bold=True, center=True, fundo=COR_H, fonte="FFFFFF")
    ws.row_dimensions[ln].height = 26


def _titulo(ws, txt, ncols, cor=COR_H):
    letra = get_column_letter(ncols)
    ws.merge_cells(f"A1:{letra}1")
    cell = ws["A1"]
    cell.value = txt
    cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=cor)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


def _aw(ws, extras=None):
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 52)
    for ci, w in (extras or {}).items():
        ws.column_dimensions[get_column_letter(ci)].width = w


def situacao(pct):
    if pct >= 75: return "Aprovado"
    if pct >= 50: return "Em Risco"
    return "Reprovado"


def exportar_formacao_excel(formacao_codigo: str, caminho: str):
    """Gera o relatório Excel completo para uma formação."""
    with get_conn() as conn:
        form = conn.execute(
            "SELECT * FROM formacoes WHERE codigo=?", (formacao_codigo,)
        ).fetchone()
        if not form:
            raise ValueError(f"Formação '{formacao_codigo}' não encontrada.")

        sessoes = conn.execute(
            "SELECT * FROM sessoes WHERE formacao_codigo=? ORDER BY data",
            (formacao_codigo,)
        ).fetchall()

        if not sessoes:
            raise ValueError("Esta formação não tem sessões importadas.")

        # Calcula total minutos do curso (soma de todas as sessões)
        total_min = sum(s["duracao_sessao_min"] for s in sessoes)
        total_sess = len(sessoes)

        # Presenças agrupadas por email (total no curso)
        alunos_raw = conn.execute("""
            SELECT p.email, p.nome,
                   SUM(p.duracao_min) AS total_min,
                   COUNT(DISTINCT s.id) AS sessoes_presente
            FROM presencas p
            JOIN sessoes s ON p.sessao_id = s.id
            WHERE s.formacao_codigo = ?
              AND p.funcao != 'Organizador'
            GROUP BY p.email
            ORDER BY total_min DESC
        """, (formacao_codigo,)).fetchall()

        # Detalhe por aluno × sessão
        detalhe_raw = conn.execute("""
            SELECT p.nome, p.email, s.data, s.titulo,
                   p.duracao_min, s.duracao_sessao_min, p.num_reconexoes,
                   s.hora_inicio_janela, s.hora_fim_janela
            FROM presencas p
            JOIN sessoes s ON p.sessao_id = s.id
            WHERE s.formacao_codigo = ?
              AND p.funcao != 'Organizador'
            ORDER BY p.nome, s.data
        """, (formacao_codigo,)).fetchall()

        # Por sessão
        sessao_stats = []
        all_emails = {r["email"] for r in alunos_raw}
        for s in sessoes:
            presentes = conn.execute("""
                SELECT COUNT(DISTINCT email) FROM presencas
                WHERE sessao_id=? AND funcao != 'Organizador'
            """, (s["id"],)).fetchone()[0]
            media = conn.execute("""
                SELECT AVG(duracao_min) FROM presencas
                WHERE sessao_id=? AND funcao != 'Organizador'
            """, (s["id"],)).fetchone()[0] or 0
            sessao_stats.append({
                "data":        s["data"],
                "titulo":      s["titulo"],
                "dur_min":     s["duracao_sessao_min"],
                "presentes":   presentes,
                "ausentes":    len(all_emails) - presentes,
                "freq_pct":    round(presentes / max(len(all_emails), 1) * 100, 1),
                "media_min":   round(media, 1),
                "janela":      f"{s['hora_inicio_janela']} – {s['hora_fim_janela']}"
                               if s["hora_inicio_janela"] else "Duração original",
            })

    wb = Workbook()
    wb.remove(wb.active)

    # ── Aba 1: Resumo por Aluno (expandido com Data e Sessão) ──────────
    ws1 = wb.create_sheet("Resumo Alunos")
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A3"

    nome_form = form["nome"] if form["nome"] != formacao_codigo else formacao_codigo
    _titulo(ws1, f"APROVEITAMENTO — {nome_form}  ({formacao_codigo})", 10)
    _hdr(ws1, 2, ["#", "Nome", "E-mail", "Data", "Sessão",
                  "Total\nSessões", "Tempo\nOnline",
                  "Frequência", "Aproveitamento", "Situação"])

    # Pré-calcula totais por aluno para preencher métricas globais
    totais_aluno = {}
    for r in alunos_raw:
        aprov = round(r["total_min"] / total_min * 100, 1) if total_min else 0
        freq  = round(r["sessoes_presente"] / total_sess * 100, 1)
        horas = f"{int(r['total_min']//60)}h {int(r['total_min']%60):02d}min"
        sit   = situacao(aprov)
        totais_aluno[r["email"]] = {
            "aprov": aprov, "freq": freq,
            "horas": horas, "sit": sit,
        }

    # Uma linha por aluno × sessão (dados do detalhe)
    ult_email = None
    for i, r in enumerate(detalhe_raw):
        ln = i + 3
        cor = COR_A if i % 2 == 0 else COR_B
        t   = totais_aluno.get(r["email"], {})
        aprov = t.get("aprov", 0)
        sit   = t.get("sit", "Reprovado")
        # Só mostra nome/email/métricas globais na primeira linha do aluno
        is_first = r["email"] != ult_email
        _c(ws1, ln, 1,  i+1,                          center=True, fundo=cor)
        _c(ws1, ln, 2,  r["nome"] if is_first else "", bold=is_first, fundo=cor)
        _c(ws1, ln, 3,  r["email"] if is_first else "",               fundo=cor)
        _c(ws1, ln, 4,  r["data"],                    center=True,  fundo=cor)
        _c(ws1, ln, 5,  r["titulo"],                                  fundo=cor)
        _c(ws1, ln, 6,  total_sess if is_first else "",center=True,  fundo=cor)
        _c(ws1, ln, 7,  t.get("horas","") if is_first else "",
                                                       center=True,  fundo=cor)
        _c(ws1, ln, 8,  t.get("freq",0)/100 if is_first else "",
                                                       center=True,  fundo=cor, fmt="0.0%")
        _c(ws1, ln, 9,  aprov/100 if is_first else "", center=True, fundo=cor, fmt="0.0%")
        _cor_sit = COR_OK if sit == "Aprovado" else (COR_WRN if sit == "Em Risco" else COR_ERR)
        _c(ws1, ln, 10, sit if is_first else "",       center=True,  fundo=_cor_sit if is_first else cor)
        ws1.row_dimensions[ln].height = 20
        ult_email = r["email"]

    ul = 2 + len(detalhe_raw)
    ws1.conditional_formatting.add(f"I3:I{ul}", DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=1, color="2E75B6"))
    _aw(ws1, {2: 24, 3: 34, 5: 40})

    # ── Aba 2: Resumo por Sessão ─────────────────────────────────────
    ws2 = wb.create_sheet("Resumo Sessoes")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"
    _titulo(ws2, f"PRESENÇAS POR SESSÃO — {nome_form}", 8, COR_H2)
    _hdr(ws2, 2, ["Data", "Sessão / Módulo", "Janela Horária",
                  "Duração (min)", "Presentes", "Ausentes",
                  "Frequência", "Média Presença (min)"])

    for i, s in enumerate(sessao_stats):
        ln = i + 3
        cor = COR_A if i % 2 == 0 else COR_B
        _c(ws2, ln, 1, s["data"],       center=True, fundo=cor)
        _c(ws2, ln, 2, s["titulo"],                  fundo=cor)
        _c(ws2, ln, 3, s["janela"],     center=True, fundo=cor)
        _c(ws2, ln, 4, int(s["dur_min"]), center=True, fundo=cor)
        _c(ws2, ln, 5, s["presentes"], center=True, fundo=cor)
        _c(ws2, ln, 6, s["ausentes"],  center=True, fundo=cor)
        _c(ws2, ln, 7, s["freq_pct"]/100, center=True, fundo=cor, fmt="0.0%")
        _c(ws2, ln, 8, s["media_min"], center=True, fundo=cor)
        ws2.row_dimensions[ln].height = 20

    ul2 = 2 + len(sessao_stats)
    ws2.conditional_formatting.add(f"G3:G{ul2}", ColorScaleRule(
        start_type="num", start_value=0,   start_color="FFC7CE",
        mid_type="num",   mid_value=0.6,   mid_color="FFEB9C",
        end_type="num",   end_value=1,     end_color="C6EFCE"))
    _aw(ws2, {2: 42, 3: 20})

    # ── Aba 3: Detalhe Aluno × Sessão ────────────────────────────────
    ws3 = wb.create_sheet("Detalhe por Aluno")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A3"
    _titulo(ws3, f"DETALHE — ALUNO × SESSÃO — {nome_form}", 7)
    _hdr(ws3, 2, ["Nome", "Data", "Sessão",
                  "Tempo Online\n(min)", "Duração Sessão\n(min)",
                  "Aproveitamento", "Reconexões"])

    ult = None
    for i, r in enumerate(detalhe_raw):
        ln = i + 3
        cor = COR_A if i % 2 == 0 else COR_B
        aprov = round(r["duracao_min"] / r["duracao_sessao_min"] * 100, 1) \
            if r["duracao_sessao_min"] else 0
        _c(ws3, ln, 1, r["nome"], bold=(r["nome"] != ult), fundo=cor)
        _c(ws3, ln, 2, r["data"],       center=True, fundo=cor)
        _c(ws3, ln, 3, r["titulo"],                  fundo=cor)
        _c(ws3, ln, 4, round(r["duracao_min"], 1), center=True, fundo=cor)
        _c(ws3, ln, 5, int(r["duracao_sessao_min"]), center=True, fundo=cor)
        _c(ws3, ln, 6, aprov/100, center=True, fundo=cor, fmt="0.0%")
        _c(ws3, ln, 7, r["num_reconexoes"], center=True, fundo=cor)
        ws3.row_dimensions[ln].height = 20
        ult = r["nome"]

    ul3 = 2 + len(detalhe_raw)
    ws3.conditional_formatting.add(f"F3:F{ul3}", ColorScaleRule(
        start_type="num", start_value=0,   start_color="FFC7CE",
        mid_type="num",   mid_value=0.6,   mid_color="FFEB9C",
        end_type="num",   end_value=1,     end_color="C6EFCE"))
    _aw(ws3, {1: 24, 3: 42})

    wb.save(caminho)
